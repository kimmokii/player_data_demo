from pathlib import Path
import calendar

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# --------------------------------------------------------------------
# Load source workbook with KPI data
# --------------------------------------------------------------------
src_path = Path.cwd() / "kpi_dashboard_data.xlsx"
wb = load_workbook(src_path)

# Expected sheets
dau_ws = wb["DAU_daily"]        # columns: day (A), dau (B)
wau_ws = wb["WAU_weekly"]       # columns: week (A), wau (B)
mau_ws = wb["MAU_monthly"]      # columns: month_yyyy_mm (A), mau (B)
rev_ws = wb["Revenue_daily"]    # columns: day (A), revenue_eur (B)
ab_ws = wb["AB_retention"]      # raw AB data


# --------------------------------------------------------------------
# 1) Build AB-retention summary: average D1 retention per variant
# --------------------------------------------------------------------
ab_ws["H1"] = "variant"
ab_ws["I1"] = "avg_d1_retention"

variants = ["control", "A", "B"]
for i, v in enumerate(variants, start=2):
    # Variant name
    ab_ws[f"H{i}"] = v
    # Average D1 retention (column H) for rows with this variant (column B)
    ab_ws[f"I{i}"] = f'=AVERAGEIF($B:$B,H{i},$H:$H)'


# --------------------------------------------------------------------
# 2) Add month name column to MAU sheet (Jan, Feb, ...)
#    A: '2025-01' style, B: MAU, C: 'Jan', 'Feb', ...
# --------------------------------------------------------------------
mau_ws["C1"] = "month_name"
for row in range(2, mau_ws.max_row + 1):
    ym = str(mau_ws[f"A{row}"].value)  # e.g. '2025-01'
    parts = ym.split("-")
    if len(parts) == 2 and parts[1].isdigit():
        month_idx = int(parts[1])
        mau_ws[f"C{row}"] = calendar.month_abbr[month_idx]  # Jan, Feb, ...
    else:
        # Fallback: just reuse the original value
        mau_ws[f"C{row}"] = ym


# --------------------------------------------------------------------
# 3) Get or create Dashboard sheet
# --------------------------------------------------------------------
if "Dashboard" in wb.sheetnames:
    dash = wb["Dashboard"]
else:
    dash = wb.create_sheet("Dashboard")

dash["A1"] = "KPI Dashboard"


# --------------------------------------------------------------------
# Helper: add a 1-series bar chart with a single colour and visible axes
# --------------------------------------------------------------------
def add_bar_chart(
    sheet,
    title,
    data_sheet,
    cat_col,
    val_col,
    pos,
    color="4472C4",
):
    """
    Create a simple bar chart:

    - Categories from column `cat_col` (rows 2..max_row).
    - Values from column `val_col` (header in row 1, data rows 2..max_row).
    - Single series with uniform colour.
    - Axis labels and tick labels forced to be visible.
    """
    max_row = data_sheet.max_row

    data_ref = Reference(data_sheet, min_col=val_col, min_row=1, max_row=max_row)
    cat_ref = Reference(data_sheet, min_col=cat_col, min_row=2, max_row=max_row)

    chart = BarChart()
    chart.title = title

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)

    chart.legend = None
    chart.varyColors = False


    # Force axis tick labels to be shown next to the axes
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"

    # Apply a single colour to the only series
    if chart.series:
        s = chart.series[0]
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color

    sheet.add_chart(chart, pos)


# --------------------------------------------------------------------
# 4) KPI charts
# --------------------------------------------------------------------

# DAU per day
add_bar_chart(
    dash,
    title="DAU",
    data_sheet=dau_ws,
    cat_col=1,
    val_col=2,
    pos="A3",
)

# WAU per week
add_bar_chart(
    dash,
    title="WAU",
    data_sheet=wau_ws,
    cat_col=1,
    val_col=2,
    pos="M3",
)

# MAU per month (x-axis shows month names, e.g. Jan, Feb, ...)
add_bar_chart(
    dash,
    title="MAU",
    data_sheet=mau_ws,
    cat_col=3,  # month_name column
    val_col=2,
    pos="A18",
)

# Revenue per day (EUR)
add_bar_chart(
    dash,
    title="Revenue per day (EUR)",
    data_sheet=rev_ws,
    cat_col=1,
    val_col=2,
    pos="M18",
    color="9E480E",
)


# --------------------------------------------------------------------
# 5) A/B test chart: one bar per variant (control, A, B)
# --------------------------------------------------------------------
ab_max_row = 1 + len(variants)

# Data: header in I1, values I2..I4
data_ref = Reference(ab_ws, min_col=9, min_row=1, max_row=ab_max_row)
# Categories: H2..H4 (control, A, B)
cat_ref = Reference(ab_ws, min_col=8, min_row=2, max_row=ab_max_row)

ab_chart = BarChart()
ab_chart.title = "Avg D1 retention by variant"
ab_chart.add_data(data_ref, titles_from_data=True)
ab_chart.set_categories(cat_ref)

# One series with three categories, different colours per bar,
# x-axis labels show which is control/A/B so legend is not necessary.
ab_chart.varyColors = True
ab_chart.legend = None

# Make sure axis labels are visible
ab_chart.x_axis.tickLblPos = "nextTo"
ab_chart.y_axis.tickLblPos = "nextTo"
ab_chart.x_axis.delete = False
ab_chart.y_axis.delete = False
ab_chart.x_axis.majorTickMark = "out"
ab_chart.y_axis.majorTickMark = "out"

dash.add_chart(ab_chart, "A33")


# --------------------------------------------------------------------
# 6) Save dashboard workbook
# --------------------------------------------------------------------
out_path = Path.cwd() / "kpi_dashboard_with_charts.xlsx"
wb.save(out_path)
print("Saved:", out_path)
